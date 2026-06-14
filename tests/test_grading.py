# tests/test_grading.py
from unittest.mock import MagicMock, patch

from langchain_core.messages import AIMessage


class TestGradingLLM:
    def test_get_llm_uses_reasoning_settings(self, monkeypatch):
        monkeypatch.setenv("NVIDIA_API_KEY", "nvapi-test")
        from ta.tools import grading
        grading._get_llm.cache_clear()
        with patch("ta.tools.grading.ChatNVIDIA") as mock_cls:
            mock_cls.return_value = MagicMock()
            grading._get_llm()
            kwargs = mock_cls.call_args.kwargs
            assert kwargs["temperature"] == 1.0
            assert kwargs["reasoning_budget"] == 16384
            assert kwargs["chat_template_kwargs"] == {"enable_thinking": False}
        grading._get_llm.cache_clear()

    def test_grading_prompt_no_legacy_toggle(self):
        from ta.tools.grading import GRADING_SYSTEM_PROMPT
        assert "detailed thinking" not in GRADING_SYSTEM_PROMPT.lower()

    def test_analyze_submission_handles_block_content(self):
        grade_json = (
            '{"criteria_scores": {"Q": 5.0}, "score": 5.0, "max_score": 5.0, '
            '"feedback_text": "ok", "inline_comments": []}'
        )
        fake_llm = MagicMock()
        fake_llm.invoke.return_value = AIMessage(
            content=[{"type": "text", "text": grade_json}]
        )
        with patch("ta.tools.grading._get_llm", return_value=fake_llm):
            from ta.tools.grading import analyze_submission
            result = analyze_submission.func(
                submission_text="print('hi')", rubric_json="[]", assignment_type="code"
            )
        assert '"score": 5.0' in result

    def test_analyze_submission_retries_on_transient_failure(self):
        import tenacity

        from ta.tools import grading
        grade_json = (
            '{"criteria_scores": {}, "score": 1.0, "max_score": 1.0, '
            '"feedback_text": "x", "inline_comments": []}'
        )
        fake_llm = MagicMock()
        fake_llm.invoke.side_effect = [
            RuntimeError("429 Too Many Requests"),
            RuntimeError("timeout"),
            AIMessage(content=grade_json),
        ]
        grading._invoke_llm.retry.wait = tenacity.wait_none()  # no sleeping in tests
        with patch("ta.tools.grading._get_llm", return_value=fake_llm):
            from ta.tools.grading import analyze_submission
            result = analyze_submission.func(
                submission_text="x", rubric_json="[]", assignment_type="code"
            )
        assert '"score": 1.0' in result
        assert fake_llm.invoke.call_count == 3


class TestPostGradeFeedback:
    def _svc_with_submission(self, with_file=True):
        sub = {"id": "sub1", "userId": "s1"}
        if with_file:
            sub["assignmentSubmission"] = {
                "attachments": [{"driveFile": {"id": "file9"}}]
            }
        svc = MagicMock()
        (svc.courses().courseWork().studentSubmissions()
         .list().execute.return_value) = {"studentSubmissions": [sub]}
        return svc

    def test_feedback_posted_to_drive_file(self):
        svc = self._svc_with_submission(with_file=True)
        drive = MagicMock()
        with patch("ta.tools.grading.interrupt", return_value=True), \
             patch("ta.tools.classroom._classroom_service", return_value=svc), \
             patch("ta.tools.drive._drive_service", return_value=drive):
            from ta.tools.grading import post_grade
            result = post_grade.func(
                course_id="c1", coursework_id="w1", student_id="s1",
                score=9.0, feedback="Buen trabajo, revisa la sección 2.",
            )
        kwargs = drive.comments().create.call_args.kwargs
        assert kwargs["fileId"] == "file9"
        assert kwargs["body"]["content"] == "Buen trabajo, revisa la sección 2."
        assert "Feedback posted" in result

    def test_no_drive_file_reports_honestly(self):
        svc = self._svc_with_submission(with_file=False)
        drive = MagicMock()
        with patch("ta.tools.grading.interrupt", return_value=True), \
             patch("ta.tools.classroom._classroom_service", return_value=svc), \
             patch("ta.tools.drive._drive_service", return_value=drive):
            from ta.tools.grading import post_grade
            result = post_grade.func(
                course_id="c1", coursework_id="w1", student_id="s1",
                score=9.0, feedback="Texto de feedback",
            )
        drive.comments().create.assert_not_called()
        assert "NOT delivered" in result
        assert "Texto de feedback" in result

    def test_post_private_comment_removed(self):
        from ta.tools import ALL_TOOLS
        assert "post_private_comment" not in [t.name for t in ALL_TOOLS]


class TestExportGrades:
    def test_exports_matrix_to_xlsx(self, tmp_path):
        svc = MagicMock()
        svc.courses().students().list().execute.return_value = {
            "students": [
                {
                    "userId": "s1",
                    "profile": {
                        "name": {"fullName": "Ana López"},
                        "emailAddress": "ana@school.mx",
                    },
                },
                {
                    "userId": "s2",
                    "profile": {
                        "name": {"fullName": "Beto Ruiz"},
                        "emailAddress": "beto@school.mx",
                    },
                },
            ]
        }
        svc.courses().courseWork().list().execute.return_value = {
            "courseWork": [{"id": "w1", "title": "HW1"}, {"id": "w2", "title": "HW2"}]
        }
        svc.courses().courseWork().studentSubmissions().list().execute.side_effect = [
            {"studentSubmissions": [
                {"userId": "s1", "assignedGrade": 9.5},
                {"userId": "s2"},  # ungraded
            ]},
            {"studentSubmissions": [
                {"userId": "s1", "assignedGrade": 8.0},
                {"userId": "s2", "assignedGrade": 10.0},
            ]},
        ]
        out = tmp_path / "grades.xlsx"
        with patch("ta.tools.classroom._classroom_service", return_value=svc):
            from ta.tools.grading import export_grades
            result = export_grades.func(course_id="c1", output_path=str(out))

        assert out.exists()
        from openpyxl import load_workbook
        ws = load_workbook(out)["Grades"]
        rows = list(ws.values)
        assert rows[0] == ("Student", "Email", "HW1", "HW2")
        assert rows[1] == ("Ana López", "ana@school.mx", 9.5, 8.0)
        assert rows[2] == ("Beto Ruiz", "beto@school.mx", None, 10.0)
        assert "2 students" in result


class TestImportGrades:
    def test_imports_known_students_and_reports_unknown(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Email", "Grade", "Feedback"])
        ws.append(["ana@school.mx", 9.5, "Bien"])
        ws.append(["beto@school.mx", 8, ""])
        ws.append(["charlie@school.mx", 7, ""])  # not enrolled
        xlsx = tmp_path / "grades.xlsx"
        wb.save(xlsx)

        svc = MagicMock()
        svc.courses().students().list().execute.return_value = {
            "students": [
                {"userId": "u-ana", "profile": {"emailAddress": "ana@school.mx"}},
                {"userId": "u-beto", "profile": {"emailAddress": "beto@school.mx"}},
            ]
        }
        sub_list = svc.courses().courseWork().studentSubmissions().list().execute
        sub_list.side_effect = [
            {"studentSubmissions": [{
                "id": "sub-ana", "userId": "u-ana",
                "assignmentSubmission": {"attachments": [{"driveFile": {"id": "f-ana"}}]},
            }]},
            {"studentSubmissions": [{"id": "sub-beto", "userId": "u-beto"}]},
        ]
        drive = MagicMock()
        captured: list[dict] = []
        with patch("ta.tools.grading.interrupt",
                   side_effect=lambda p: captured.append(p) or True), \
             patch("ta.tools.classroom._classroom_service", return_value=svc), \
             patch("ta.tools.drive._drive_service", return_value=drive):
            from ta.tools.grading import import_grades
            result = import_grades.func(
                course_id="c1", coursework_id="w1", xlsx_path=str(xlsx)
            )

        assert "Post 2 grades" in captured[0]["details"]
        assert "Imported 2 grades" in result
        assert "charlie@school.mx" in result and "not enrolled" in result
        patch_call = svc.courses().courseWork().studentSubmissions().patch
        assert patch_call.call_count == 2
        assert drive.comments().create.call_count == 1  # only Ana had feedback+file
